"""Routes for Shared Services, DC-scoped Application Presences, and the
multi-DC Rule Request / fan-out engine.

All endpoints are additive (do not break existing reference routes).
They live under /api/reference/shared-services and /api/rules/*.
"""
from __future__ import annotations

from typing import Any

from fastapi import APIRouter, HTTPException

from app.database import (
    classify_ip,
    classify_ips,
    create_group_change_request,
    create_port,
    create_rule_request,
    create_shared_service,
    delete_app_presence,
    delete_port,
    delete_shared_service,
    delete_shared_service_presence,
    get_app_presences,
    get_group_change_artifacts,
    get_group_change_request,
    get_occupants,
    get_rule_request,
    get_rule_requests,
    get_shared_service,
    get_shared_service_presences,
    get_shared_services,
    ingest_app_members,
    itsm_webhook_dispatch,
    list_group_change_requests,
    list_ports,
    poll_itsm_connectors_once,
    preview_rule_expansion,
    refresh_group_change_external_status,
    set_group_change_request_status,
    set_rule_request_status,
    submit_group_change_request_to_itsm,
    update_port,
    update_shared_service,
    upsert_app_presence,
    upsert_shared_service_presence,
)

router = APIRouter(tags=["Shared Services & Multi-DC"])


# ---- Shared Services ----

@router.get("/api/reference/shared-services")
async def list_shared_services(
    environment: str | None = None,
    category: str | None = None,
    q: str | None = None,
    team: str | None = None,
) -> list[dict[str, Any]]:
    items = await get_shared_services()
    if environment:
        items = [s for s in items
                 if environment in (s.get("environments") or ["Production"])]
    if category:
        items = [s for s in items if s.get("category") == category]
    if q:
        ql = q.lower()
        items = [
            s for s in items
            if ql in str(s.get("service_id", "")).lower()
            or ql in str(s.get("name", "")).lower()
            or ql in str(s.get("description", "")).lower()
        ]
    # Team-scoped visibility. SNS = god view (sees everything).
    if team and team.strip().upper() != "SNS":
        t = team.strip().lower()
        items = [s for s in items
                 if str(s.get("owner_team", "")).strip().lower() == t]
    return items


@router.post("/api/reference/shared-services")
async def create_shared_service_route(payload: dict[str, Any]) -> dict[str, Any]:
    if not payload.get("service_id"):
        raise HTTPException(400, "service_id is required")
    if not payload.get("name"):
        raise HTTPException(400, "name is required")
    return await create_shared_service(payload)


@router.get("/api/reference/shared-services/{service_id}")
async def get_shared_service_route(service_id: str) -> dict[str, Any]:
    s = await get_shared_service(service_id)
    if not s:
        raise HTTPException(404, f"Shared service {service_id} not found")
    return s


@router.put("/api/reference/shared-services/{service_id}")
async def update_shared_service_route(service_id: str,
                                       payload: dict[str, Any]) -> dict[str, Any]:
    s = await update_shared_service(service_id, payload)
    if not s:
        raise HTTPException(404, f"Shared service {service_id} not found")
    return s


@router.delete("/api/reference/shared-services/{service_id}")
async def delete_shared_service_route(service_id: str) -> dict[str, str]:
    ok = await delete_shared_service(service_id)
    if not ok:
        raise HTTPException(404, f"Shared service {service_id} not found")
    return {"status": "deleted", "service_id": service_id}


# ---- Shared Service Presences ----

@router.get("/api/reference/shared-services/{service_id}/presences")
async def list_shared_service_presences(service_id: str,
                                         environment: str | None = None,
                                         dc_id: str | None = None) -> list[dict[str, Any]]:
    items = await get_shared_service_presences(service_id)
    if environment:
        items = [p for p in items if p.get("environment") == environment]
    if dc_id:
        items = [p for p in items if p.get("dc_id") == dc_id]
    return items


@router.post("/api/reference/shared-services/{service_id}/presences")
async def upsert_shared_service_presence_route(service_id: str,
                                                 payload: dict[str, Any]) -> dict[str, Any]:
    payload = dict(payload)
    payload["service_id"] = service_id
    for required in ("dc_id", "environment", "nh_id", "sz_code"):
        if not payload.get(required):
            raise HTTPException(400, f"{required} is required on presence")
    return await upsert_shared_service_presence(payload)


@router.delete("/api/reference/shared-services/{service_id}/presences")
async def delete_shared_service_presence_route(service_id: str, dc_id: str,
                                                 environment: str, nh_id: str,
                                                 sz_code: str) -> dict[str, str]:
    ok = await delete_shared_service_presence(
        service_id, dc_id, environment, nh_id, sz_code)
    if not ok:
        raise HTTPException(404, "Presence not found")
    return {"status": "deleted"}


# ---- App Presences (DC-scoped) ----

@router.get("/api/reference/app-presences")
async def list_app_presences(app: str | None = None,
                              environment: str | None = None,
                              dc_id: str | None = None) -> list[dict[str, Any]]:
    items = await get_app_presences(app)
    if environment:
        items = [p for p in items if p.get("environment") == environment]
    if dc_id:
        items = [p for p in items if p.get("dc_id") == dc_id]
    return items


@router.post("/api/reference/app-presences")
async def upsert_app_presence_route(payload: dict[str, Any]) -> dict[str, Any]:
    for required in ("app_distributed_id", "dc_id", "environment",
                     "nh_id", "sz_code"):
        if not payload.get(required):
            raise HTTPException(400, f"{required} is required on app presence")
    return await upsert_app_presence(payload)


@router.delete("/api/reference/app-presences")
async def delete_app_presence_route(app_distributed_id: str, dc_id: str,
                                      environment: str, nh_id: str,
                                      sz_code: str) -> dict[str, str]:
    ok = await delete_app_presence(
        app_distributed_id, dc_id, environment, nh_id, sz_code)
    if not ok:
        raise HTTPException(404, "App presence not found")
    return {"status": "deleted"}


# ---- Bidirectional IP ↔ (DC, NH, SZ) classifier / occupants ----

@router.post("/api/reference/classify-ip")
async def classify_ip_route(payload: dict[str, Any]) -> dict[str, Any]:
    """Classify a single IP/CIDR/range into (DC, NH, SZ).

    Body: {"ip": "10.1.2.3", "dc_hint": "DC1" (optional)}
    """
    ip = str(payload.get("ip", "")).strip()
    if not ip:
        raise HTTPException(400, "ip is required")
    return await classify_ip(ip, str(payload.get("dc_hint", "")))


@router.post("/api/reference/classify-ips")
async def classify_ips_route(payload: dict[str, Any]) -> list[dict[str, Any]]:
    """Batch classifier.

    Body: {"ips": ["10.1.2.3", "10.2.0.0/24"], "dc_hint": "..." (optional)}
    """
    ips = payload.get("ips") or []
    if not isinstance(ips, list):
        raise HTTPException(400, "ips must be a list")
    return await classify_ips([str(x) for x in ips], str(payload.get("dc_hint", "")))


@router.get("/api/reference/occupants")
async def occupants_route(dc: str = "", nh: str = "", sz: str = "",
                           environment: str = "") -> dict[str, Any]:
    """Reverse lookup: apps + shared services present in a (DC, NH, SZ) cell.

    Any of dc/nh/sz/environment may be empty (wildcard on that axis).
    """
    return await get_occupants(dc=dc, nh=nh, sz=sz, environment=environment)


@router.post("/api/reference/applications/{app_distributed_id}/ingest-members")
async def ingest_app_members_route(app_distributed_id: str,
                                     payload: dict[str, Any]) -> dict[str, Any]:
    """Classify IPs/CIDRs for an app and auto-materialize per-(DC, NH, SZ)
    presences + derived egress/ingress groups.

    Body: {
      "environment": "Production",
      "direction": "egress" | "ingress",
      "dc_hint": "DC1" (optional),
      "members": [{"kind": "ip"|"cidr"|"range"|"subnet", "value": "..."}, ...]
    }
    """
    env = str(payload.get("environment", "Production"))
    direction = str(payload.get("direction", "egress")).lower()
    members = payload.get("members") or []
    if not isinstance(members, list):
        raise HTTPException(400, "members must be a list")
    return await ingest_app_members(
        app_distributed_id=app_distributed_id,
        environment=env,
        members=[dict(m) for m in members],
        direction=direction,
        dc_hint=str(payload.get("dc_hint", "")),
        default_has_ingress=bool(payload.get("has_ingress", False)),
    )


# ---- Rule Requests (multi-DC fan-out) ----

@router.post("/api/rules/preview-expansion")
async def preview_expansion_route(payload: dict[str, Any]) -> dict[str, Any]:
    return await preview_rule_expansion(payload)


@router.get("/api/rules/requests")
async def list_rule_requests(environment: str | None = None,
                              status: str | None = None,
                              team: str | None = None) -> list[dict[str, Any]]:
    """List rule requests. Team filter scopes results to requests raised
    for apps/services owned by `team`. SNS sees every request (global
    reviewer/approver)."""
    items = await get_rule_requests()
    if environment:
        items = [r for r in items if r.get("environment") == environment]
    if status:
        items = [r for r in items if r.get("status") == status]
    if team and team.strip().upper() != "SNS":
        t = team.strip().lower()
        items = [r for r in items
                 if str(r.get("owner_team", "")).strip().lower() == t]
    return items


@router.post("/api/rules/requests")
async def create_rule_request_route(payload: dict[str, Any]) -> dict[str, Any]:
    src_kind = (payload.get("source_kind") or "app").lower()
    src_ref = payload.get("source_ref") or payload.get("application_ref")
    if not src_ref:
        raise HTTPException(
            400,
            "source_ref (or legacy application_ref) is required",
        )
    if not payload.get("destination_kind"):
        raise HTTPException(400, "destination_kind is required")
    payload["source_kind"] = src_kind
    payload["source_ref"] = src_ref
    return await create_rule_request(payload)


@router.get("/api/rules/requests/{request_id}")
async def get_rule_request_route(request_id: str) -> dict[str, Any]:
    r = await get_rule_request(request_id)
    if not r:
        raise HTTPException(404, "Rule request not found")
    return r


@router.put("/api/rules/requests/{request_id}/status")
async def set_rule_request_status_route(request_id: str,
                                          payload: dict[str, Any]) -> dict[str, Any]:
    status = payload.get("status")
    note = payload.get("note")
    if not status:
        raise HTTPException(400, "status is required")
    r = await set_rule_request_status(request_id, status, note)
    if not r:
        raise HTTPException(404, "Rule request not found")
    return r


# ---- Port / Service Catalog ----

@router.get("/api/reference/ports")
async def list_ports_route() -> list[dict[str, Any]]:
    """Return the catalog of well-known + custom ports used by the Rule Builder."""
    return await list_ports()


@router.post("/api/reference/ports")
async def create_port_route(payload: dict[str, Any]) -> dict[str, Any]:
    if not payload.get("name") and not payload.get("port_id"):
        raise HTTPException(400, "name or port_id is required")
    try:
        int(payload.get("port") or 0)
    except (TypeError, ValueError):
        raise HTTPException(400, "port must be an integer")
    return await create_port(payload)


@router.put("/api/reference/ports/{port_id}")
async def update_port_route(port_id: str, payload: dict[str, Any]) -> dict[str, Any]:
    r = await update_port(port_id, payload)
    if not r:
        raise HTTPException(404, "Port not found")
    return r


@router.delete("/api/reference/ports/{port_id}")
async def delete_port_route(port_id: str) -> dict[str, Any]:
    ok = await delete_port(port_id)
    if not ok:
        raise HTTPException(404, "Port not found")
    return {"deleted": True, "port_id": port_id}


# ============================================================
# Phase A: Validation, Artifacts, ITSM Connector, Migration
# ============================================================
from app.database import (  # noqa: E402
    delete_birthright_rule,
    delete_itsm_connector,
    evaluate_dedup,
    evaluate_birthright,
    get_request_artifacts,
    get_security_zones,
    list_birthright_rules,
    list_itsm_connectors,
    normalize_legacy_rule,
    normalize_legacy_rules_bulk,
    refresh_request_external_status,
    submit_request_to_itsm,
    update_security_zone,
    upsert_birthright_rule,
    upsert_itsm_connector,
)


# ---- Validation: dedup + birthright on demand ----

@router.post("/api/rules/validate")
async def validate_rule(payload: dict[str, Any]) -> dict[str, Any]:
    """Run the same dedup + birthright engine that gates submit, without
    persisting anything. Returns ``{ dedup, birthright, block_submit }``."""

    # If the caller passes a payload that *could* be expanded, expand it
    # first so the dedup engine can compare full PhysicalRules. Otherwise
    # treat the payload as an already-expanded list.
    if payload.get("source_ref") or payload.get("application_ref"):
        from app.database import preview_rule_expansion
        preview = await preview_rule_expansion(payload)
        return {
            "dedup": preview.get("dedup"),
            "birthright": preview.get("birthright"),
            "block_submit": preview.get("block_submit", False),
            "physical_rules": preview.get("physical_rules", []),
            "warnings": preview.get("warnings", []),
        }
    dedup = evaluate_dedup(payload)
    birthright = evaluate_birthright(payload)
    return {
        "dedup": dedup,
        "birthright": birthright,
        "block_submit": bool(dedup.get("block") or birthright.get("covered")),
    }


# ---- Birthright Rule Registry ----

@router.get("/api/birthright-rules")
async def list_birthright_rules_route() -> list[dict[str, Any]]:
    return await list_birthright_rules()


@router.post("/api/birthright-rules")
async def upsert_birthright_rule_route(payload: dict[str, Any]) -> dict[str, Any]:
    return await upsert_birthright_rule(payload)


@router.delete("/api/birthright-rules/{birthright_id}")
async def delete_birthright_rule_route(birthright_id: str) -> dict[str, Any]:
    ok = await delete_birthright_rule(birthright_id)
    if not ok:
        raise HTTPException(404, "Birthright rule not found")
    return {"status": "deleted"}


# ---- Security Zone naming_mode toggle ----

@router.put("/api/reference/security-zones/{code}/naming-mode")
async def set_sz_naming_mode(code: str, payload: dict[str, Any]) -> dict[str, Any]:
    mode = str(payload.get("naming_mode") or "").lower().strip()
    if mode not in ("app_scoped", "zone_scoped"):
        raise HTTPException(400, "naming_mode must be app_scoped or zone_scoped")
    r = await update_security_zone(code, {"naming_mode": mode})
    if not r:
        raise HTTPException(404, "Security zone not found")
    return r


@router.get("/api/reference/security-zones-with-mode")
async def list_security_zones_with_mode() -> list[dict[str, Any]]:
    """Return the SZ catalog including ``naming_mode`` for the Settings UI."""

    return await get_security_zones()


# ---- Deployment Artifacts ----

@router.get("/api/rules/requests/{request_id}/artifacts")
async def request_artifacts_route(request_id: str) -> dict[str, Any]:
    artifacts = await get_request_artifacts(request_id)
    if artifacts is None:
        raise HTTPException(404, "Rule request not found")
    return artifacts


@router.get("/api/rules/requests/{request_id}/artifacts/manifest.json")
async def request_artifact_manifest_json(request_id: str) -> dict[str, Any]:
    artifacts = await get_request_artifacts(request_id)
    if artifacts is None:
        raise HTTPException(404, "Rule request not found")
    return artifacts["manifest"]


@router.get("/api/rules/requests/{request_id}/artifacts/manifest.xlsx")
async def request_artifact_manifest_xlsx(request_id: str):
    """Stream a real .xlsx file for ServiceNow / SNS attachments."""

    from io import BytesIO
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse

    artifacts = await get_request_artifacts(request_id)
    if artifacts is None:
        raise HTTPException(404, "Rule request not found")
    sheets = artifacts["xlsx_sheets"]
    wb = Workbook()
    # Remove the default sheet then add ours in order.
    default = wb.active
    wb.remove(default)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name[:31] or "Sheet")
        for row in rows:
            ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                f'attachment; filename="{request_id}-manifest.xlsx"'
            )
        },
    )


@router.get("/api/rules/requests/{request_id}/artifacts/device.{vendor}")
async def request_artifact_vendor_config(request_id: str, vendor: str):
    from fastapi.responses import PlainTextResponse
    artifacts = await get_request_artifacts(request_id)
    if artifacts is None:
        raise HTTPException(404, "Rule request not found")
    vendor_l = vendor.lower()
    cfg = artifacts["vendor_configs"].get(vendor_l)
    if cfg is None:
        raise HTTPException(
            404, f"Unknown vendor '{vendor}'. Try one of: " +
            ", ".join(artifacts["vendor_configs"].keys()))
    return PlainTextResponse(
        cfg,
        media_type="text/plain",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{request_id}-{vendor_l}.txt"'
            )
        },
    )


def _build_bundle_zip(artifacts: dict[str, Any], request_id: str) -> bytes:
    """Pack JSON manifest + XLSX + every vendor config into a single zip
    so SNS can grab everything in one click and attach it to a CR even
    while the ITSM integration is still being wired up."""

    import io
    import json as _json
    import zipfile
    from openpyxl import Workbook

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # JSON manifest
        zf.writestr(
            f"{request_id}/manifest.json",
            _json.dumps(artifacts["manifest"], indent=2, default=str),
        )
        # XLSX manifest
        wb = Workbook()
        default = wb.active
        wb.remove(default)
        for name, rows in artifacts["xlsx_sheets"].items():
            ws = wb.create_sheet(title=(name or "Sheet")[:31])
            for row in rows:
                ws.append(row)
        xlsx_buf = io.BytesIO()
        wb.save(xlsx_buf)
        zf.writestr(f"{request_id}/manifest.xlsx", xlsx_buf.getvalue())
        # Each vendor config as a .txt
        for vendor_l, cfg in (artifacts.get("vendor_configs") or {}).items():
            zf.writestr(f"{request_id}/device-{vendor_l}.txt", cfg or "")
        # README so SNS / consumers know what's in the bundle
        readme = (
            f"NGDC Firewall Studio — Rule Request {request_id} export bundle\n"
            f"================================================\n"
            f"manifest.json     Vendor-neutral deployment manifest\n"
            f"manifest.xlsx     Same data flattened to spreadsheet rows\n"
            f"device-*.txt      Vendor-compiled configs (groups + rules)\n"
            f"\n"
            f"Use any of the above to deploy manually if the ITSM "
            f"integration is not yet wired up.\n"
        )
        zf.writestr(f"{request_id}/README.txt", readme)
    return buf.getvalue()


@router.get("/api/rules/requests/{request_id}/artifacts/bundle.zip")
async def request_artifact_bundle_zip(request_id: str):
    """Single-click "Download All" — every artifact format zipped together
    so it can be manually attached to a ServiceNow CR / emailed / handed
    off to SNS while the automated ITSM integration is being set up."""

    from fastapi.responses import StreamingResponse
    from io import BytesIO

    artifacts = await get_request_artifacts(request_id)
    if artifacts is None:
        raise HTTPException(404, "Rule request not found")
    payload = _build_bundle_zip(artifacts, request_id)
    return StreamingResponse(
        BytesIO(payload),
        media_type="application/zip",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{request_id}-bundle.zip"'
            )
        },
    )


@router.post("/api/rules/requests/artifacts/bulk-bundle.zip")
async def request_artifact_bulk_bundle_zip(payload: dict[str, Any]):
    """Bulk export — pack one folder per requested rule into a single
    zip. Body: ``{"request_ids": ["RQ-1", "RQ-2", ...]}``. Skips IDs
    that don't resolve to a rule request (returns the rest)."""

    from fastapi.responses import StreamingResponse
    from io import BytesIO
    import zipfile
    import json as _json
    from openpyxl import Workbook

    ids = payload.get("request_ids") or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, "request_ids must be a non-empty list")

    buf = BytesIO()
    included: list[str] = []
    skipped: list[str] = []
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for rid in ids:
            rid_s = str(rid)
            arts = await get_request_artifacts(rid_s)
            if arts is None:
                skipped.append(rid_s)
                continue
            included.append(rid_s)
            # JSON
            zf.writestr(
                f"{rid_s}/manifest.json",
                _json.dumps(arts["manifest"], indent=2, default=str),
            )
            # XLSX
            wb = Workbook()
            default = wb.active
            wb.remove(default)
            for name, rows in arts["xlsx_sheets"].items():
                ws = wb.create_sheet(title=(name or "Sheet")[:31])
                for row in rows:
                    ws.append(row)
            xlsx_buf = BytesIO()
            wb.save(xlsx_buf)
            zf.writestr(f"{rid_s}/manifest.xlsx", xlsx_buf.getvalue())
            # Vendor configs
            for vendor_l, cfg in (arts.get("vendor_configs") or {}).items():
                zf.writestr(f"{rid_s}/device-{vendor_l}.txt", cfg or "")
        # Index file at the root.
        index = {
            "exported_request_ids": included,
            "skipped_request_ids": skipped,
            "format_legend": {
                "manifest.json": "Vendor-neutral deployment manifest",
                "manifest.xlsx": "Same data flattened to spreadsheet rows",
                "device-*.txt": "Vendor-compiled configs (groups + rules)",
            },
        }
        zf.writestr("INDEX.json", _json.dumps(index, indent=2))
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={
            "Content-Disposition": (
                'attachment; filename="rule-requests-bulk-export.zip"'
            )
        },
    )


# ---- ITSM Connector Framework ----

@router.get("/api/itsm/connectors")
async def list_itsm_connectors_route() -> list[dict[str, Any]]:
    return await list_itsm_connectors()


@router.post("/api/itsm/connectors")
async def upsert_itsm_connector_route(payload: dict[str, Any]) -> dict[str, Any]:
    if not payload.get("kind"):
        raise HTTPException(400, "kind is required (servicenow | generic_rest | internal)")
    if not payload.get("endpoint_url"):
        raise HTTPException(400, "endpoint_url is required")
    return await upsert_itsm_connector(payload)


@router.delete("/api/itsm/connectors/{connector_id}")
async def delete_itsm_connector_route(connector_id: str) -> dict[str, Any]:
    ok = await delete_itsm_connector(connector_id)
    if not ok:
        raise HTTPException(404, "Connector not found")
    return {"status": "deleted"}


@router.post("/api/rules/requests/{request_id}/submit-itsm")
async def submit_itsm_route(request_id: str,
                             payload: dict[str, Any] | None = None) -> dict[str, Any]:
    payload = payload or {}
    res = await submit_request_to_itsm(
        request_id, payload.get("connector_id"))
    if res is None:
        raise HTTPException(404, "Rule request not found")
    return res


@router.post("/api/rules/requests/{request_id}/refresh-external-status")
async def refresh_external_status_route(request_id: str) -> dict[str, Any]:
    res = await refresh_request_external_status(request_id)
    if res is None:
        raise HTTPException(404, "Rule request not found")
    return res


# ---- Migration Normalizer ----

@router.post("/api/migration/normalize")
async def normalize_legacy_rule_route(payload: dict[str, Any]) -> dict[str, Any]:
    return await normalize_legacy_rule(payload)


@router.post("/api/migration/normalize-bulk")
async def normalize_legacy_rules_bulk_route(payload: dict[str, Any]) -> dict[str, Any]:
    rules = payload.get("rules") or []
    if not isinstance(rules, list):
        raise HTTPException(400, "rules must be a list")
    return await normalize_legacy_rules_bulk([dict(r) for r in rules])


# ============================================================
# Group Change Requests (standalone group create / modify / delete)
# ============================================================

@router.get("/api/groups/change-requests")
async def list_group_change_requests_route(
    environment: str | None = None,
    status: str | None = None,
    team: str | None = None,
) -> list[dict[str, Any]]:
    return await list_group_change_requests(
        environment=environment, status=status, team=team)


@router.post("/api/groups/change-requests")
async def create_group_change_request_route(
    payload: dict[str, Any],
) -> dict[str, Any]:
    try:
        return await create_group_change_request(payload)
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.get("/api/groups/change-requests/{request_id}")
async def get_group_change_request_route(request_id: str) -> dict[str, Any]:
    res = await get_group_change_request(request_id)
    if res is None:
        raise HTTPException(404, "Group change request not found")
    return res


@router.patch("/api/groups/change-requests/{request_id}/status")
async def set_group_change_request_status_route(
    request_id: str, payload: dict[str, Any],
) -> dict[str, Any]:
    status = payload.get("status")
    note = payload.get("note")
    if not status:
        raise HTTPException(400, "status is required")
    try:
        res = await set_group_change_request_status(request_id, status, note)
    except ValueError as e:
        raise HTTPException(400, str(e))
    if res is None:
        raise HTTPException(404, "Group change request not found")
    return res


@router.get("/api/groups/change-requests/{request_id}/artifacts")
async def group_change_artifacts_route(request_id: str) -> dict[str, Any]:
    res = await get_group_change_artifacts(request_id)
    if res is None:
        raise HTTPException(404, "Group change request not found")
    return res


@router.get("/api/groups/change-requests/{request_id}/artifacts/manifest.json")
async def group_change_artifact_manifest_json(request_id: str) -> dict[str, Any]:
    arts = await get_group_change_artifacts(request_id)
    if arts is None:
        raise HTTPException(404, "Group change request not found")
    return arts["manifest"]


@router.get("/api/groups/change-requests/{request_id}/artifacts/manifest.xlsx")
async def group_change_artifact_manifest_xlsx(request_id: str):
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    from openpyxl import Workbook

    arts = await get_group_change_artifacts(request_id)
    if arts is None:
        raise HTTPException(404, "Group change request not found")
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for name, rows in arts["xlsx_sheets"].items():
        ws = wb.create_sheet(title=(name or "Sheet")[:31])
        for row in rows:
            ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{request_id}-group-change.xlsx"'
            )
        },
    )


@router.get("/api/groups/change-requests/{request_id}/artifacts/device-{vendor}.txt")
async def group_change_artifact_vendor_config(request_id: str, vendor: str):
    from fastapi.responses import PlainTextResponse

    arts = await get_group_change_artifacts(request_id)
    if arts is None:
        raise HTTPException(404, "Group change request not found")
    cfg = (arts.get("vendor_configs") or {}).get(vendor.lower())
    if cfg is None:
        raise HTTPException(404, f"Unknown vendor {vendor!r}")
    return PlainTextResponse(
        cfg,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{request_id}-{vendor}.txt"'
            )
        },
    )


@router.get("/api/groups/change-requests/{request_id}/artifacts/bundle.zip")
async def group_change_artifact_bundle_zip(request_id: str):
    from fastapi.responses import StreamingResponse
    from io import BytesIO

    arts = await get_group_change_artifacts(request_id)
    if arts is None:
        raise HTTPException(404, "Group change request not found")
    payload = _build_bundle_zip(arts, request_id)
    return StreamingResponse(
        BytesIO(payload),
        media_type="application/zip",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{request_id}-group-bundle.zip"'
            )
        },
    )


@router.post("/api/groups/change-requests/artifacts/bulk-bundle.zip")
async def group_change_bulk_bundle_zip(payload: dict[str, Any]):
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    import zipfile
    import json as _json
    from openpyxl import Workbook

    ids = payload.get("request_ids") or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, "request_ids must be a non-empty list")

    buf = BytesIO()
    included: list[str] = []
    skipped: list[str] = []
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for rid in ids:
            rid_s = str(rid)
            arts = await get_group_change_artifacts(rid_s)
            if arts is None:
                skipped.append(rid_s)
                continue
            included.append(rid_s)
            zf.writestr(
                f"{rid_s}/manifest.json",
                _json.dumps(arts["manifest"], indent=2, default=str),
            )
            wb = Workbook()
            default = wb.active
            wb.remove(default)
            for name, rows in arts["xlsx_sheets"].items():
                ws = wb.create_sheet(title=(name or "Sheet")[:31])
                for row in rows:
                    ws.append(row)
            xlsx_buf = BytesIO()
            wb.save(xlsx_buf)
            zf.writestr(f"{rid_s}/manifest.xlsx", xlsx_buf.getvalue())
            for vendor_l, cfg in (arts.get("vendor_configs") or {}).items():
                zf.writestr(f"{rid_s}/device-{vendor_l}.txt", cfg or "")
        index = {
            "exported_request_ids": included,
            "skipped_request_ids": skipped,
            "kind": "group_change_request",
        }
        zf.writestr("INDEX.json", _json.dumps(index, indent=2))
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={
            "Content-Disposition": (
                'attachment; filename="group-change-requests-bulk-export.zip"'
            )
        },
    )


@router.post("/api/groups/change-requests/{request_id}/submit-itsm")
async def submit_group_change_to_itsm_route(
    request_id: str,
    payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    payload = payload or {}
    try:
        res = await submit_group_change_request_to_itsm(
            request_id, payload.get("connector_id"))
    except ValueError as e:
        raise HTTPException(400, str(e))
    if res is None:
        raise HTTPException(404, "Group change request not found")
    return res


@router.post("/api/groups/change-requests/{request_id}/refresh-external-status")
async def refresh_group_change_status_route(request_id: str) -> dict[str, Any]:
    res = await refresh_group_change_external_status(request_id)
    if res is None:
        raise HTTPException(404, "Group change request not found")
    return res


# ============================================================
# Phase B — ITSM polling worker + webhook receiver
# ============================================================

@router.post("/api/itsm/poll")
async def itsm_poll_once_route() -> dict[str, Any]:
    """Trigger one polling pass over every ITSM connector. Returns the
    list of state transitions applied. Safe to invoke from a cron / k8s
    CronJob / Celery beat — keeps things idempotent."""
    return await poll_itsm_connectors_once()


@router.post("/api/itsm/webhook/{connector_id}")
async def itsm_webhook_route(
    connector_id: str, payload: dict[str, Any],
) -> dict[str, Any]:
    """Receive a push from ServiceNow Business Rules / generic webhook /
    internal tool. Body must include ``external_ticket_id`` (or
    ``ticket_id`` / ``number``) and ``state`` (or ``status``). Matches
    the ticket back to a RuleRequest or GroupChangeRequest and applies
    the transition automatically."""
    try:
        return await itsm_webhook_dispatch(
            connector_id=connector_id, payload=payload)
    except ValueError as e:
        raise HTTPException(400, str(e))
