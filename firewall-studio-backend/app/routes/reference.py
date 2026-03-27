from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from app.database import (
    get_ngdc_datacenters, get_security_zones, get_predefined_destinations,
    get_neighbourhoods, get_legacy_datacenters, get_applications,
    get_environments, get_chg_requests, get_naming_standards, get_org_config,
    get_policy_matrix, get_heritage_dc_matrix, get_ngdc_prod_matrix, get_nonprod_matrix,
    get_rules, get_legacy_rules, update_legacy_rule, delete_legacy_rule, clear_all_legacy_rules,
    import_legacy_rules, get_migration_history, migrate_rule_to_ngdc,
    create_migration_review,
    create_neighbourhood, update_neighbourhood, delete_neighbourhood,
    create_security_zone, update_security_zone, delete_security_zone,
    create_application, update_application, delete_application,
    create_datacenter, update_datacenter, delete_datacenter,
    create_predefined_destination, update_predefined_destination, delete_predefined_destination,
    create_environment, delete_environment,
    update_org_config, update_naming_standards,
    create_policy_entry, delete_policy_entry,
    create_chg_request,
    get_groups, get_group, create_group, update_group, delete_group,
    add_group_member, remove_group_member,
    create_rule_modification, get_rule_modifications, approve_rule_modification, reject_rule_modification,
    compile_legacy_rule, validate_birthright,
    get_birthright_matrix, update_birthright_matrix, add_birthright_entry, delete_birthright_entry,
    get_ngdc_recommendations,
)
import io
import json as json_lib
import csv
import openpyxl
import zipfile
from fastapi.responses import StreamingResponse
from app.services.naming_standards import (
    validate_name, generate_group_name, generate_server_name,
    generate_subnet_name, suggest_standard_name, determine_security_zone,
    get_naming_standards_info,
)

router = APIRouter(prefix="/api/reference", tags=["Reference Data"])


# ---- Data Mode Toggle (Seed vs Live) ----

@router.get("/data-mode")
async def get_data_mode_endpoint():
    """Return the current data mode: 'seed' or 'live'."""
    from app.database import get_data_mode
    return {"mode": get_data_mode()}


@router.post("/data-mode")
async def set_data_mode_endpoint(data: dict):
    """Switch data mode. 'seed' = test/seed data, 'live' = real imported data."""
    from app.database import set_data_mode
    mode = data.get("mode", "")
    if mode not in ("seed", "live"):
        raise HTTPException(status_code=400, detail="mode must be 'seed' or 'live'")
    result = set_data_mode(mode)
    return {"mode": result}


@router.post("/data-mode/reset-seed")
async def reset_seed_data():
    """Re-seed the seed data directory (does NOT touch live data)."""
    from app.database import seed_database, set_data_mode, get_data_mode
    prev = get_data_mode()
    set_data_mode("seed")
    await seed_database()
    set_data_mode(prev)
    return {"message": "Seed data reset successfully", "current_mode": prev}


def _parse_excel_bytes(contents: bytes, filename: str = "") -> list[tuple]:
    """Parse Excel/CSV file bytes into a list of row tuples.
    Tries multiple strategies in order: CSV, openpyxl (all sheets), xlrd, HTML-table, ZIP-extract.
    Returns list of tuples where first tuple is headers."""
    import logging
    logger = logging.getLogger("excel_import")
    logger.info(f"Received file: {len(contents)} bytes, name={filename}, first 8 bytes: {contents[:8].hex() if contents else 'empty'}")

    errors: list[str] = []

    # Strategy 0: CSV file (explicit .csv extension or BOM/text heuristic)
    if filename.lower().endswith(".csv") or (not filename and contents[:3] in (b'\xef\xbb\xbf', b'App')):
        try:
            # Try multiple encodings
            for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
                try:
                    text = contents.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                text = contents.decode("utf-8", errors="replace")
            reader = csv.reader(io.StringIO(text))
            rows = [tuple(row) for row in reader if any(cell.strip() for cell in row)]
            if rows:
                logger.info(f"CSV success: {len(rows)} rows (encoding={encoding})")
                return rows
        except Exception as exc:
            errors.append(f"csv: {exc}")
            logger.warning(f"CSV parse failed: {exc}")

    # Strategy 1: openpyxl for .xlsx (Office Open XML)
    # Try read_only=True first for performance, fall back to read_only=False for protected files
    for read_only in (True, False):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=read_only, data_only=True)
            # Try active sheet first, then iterate all sheets to find the one with most rows
            best_rows: list[tuple] = []
            for ws in wb.worksheets:
                try:
                    sheet_rows = list(ws.iter_rows(values_only=True))
                    if len(sheet_rows) > len(best_rows):
                        best_rows = sheet_rows
                except Exception:
                    continue
            wb.close()
            if best_rows:
                logger.info(f"openpyxl success (read_only={read_only}): {len(best_rows)} rows")
                return best_rows
        except Exception as exc:
            errors.append(f"xlsx(read_only={read_only}): {exc}")
            logger.warning(f"openpyxl (read_only={read_only}) failed: {exc}")

    # Strategy 2: xlrd for old .xls (Excel 97-2003 binary BIFF format)
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=contents)
        # Find the sheet with the most rows
        best_sheet = wb.sheet_by_index(0)
        for i in range(wb.nsheets):
            s = wb.sheet_by_index(i)
            if s.nrows > best_sheet.nrows:
                best_sheet = s
        rows_list: list[tuple] = []
        for row_idx in range(best_sheet.nrows):
            rows_list.append(tuple(best_sheet.cell_value(row_idx, col_idx) for col_idx in range(best_sheet.ncols)))
        if rows_list:
            logger.info(f"xlrd success: {len(rows_list)} rows from sheet '{best_sheet.name}'")
            return rows_list
    except Exception as exc:
        errors.append(f"xls: {exc}")
        logger.warning(f"xlrd failed: {exc}")

    # Strategy 3: HTML table disguised as .xlsx (some tools export this way)
    try:
        text = contents.decode("utf-8", errors="ignore")
        if "<table" in text.lower() or "<html" in text.lower():
            from html.parser import HTMLParser
            html_rows: list[tuple] = []
            current_row: list[str] = []
            in_cell = False
            cell_text = ""

            class TableParser(HTMLParser):
                def handle_starttag(self, tag: str, attrs: list) -> None:
                    nonlocal in_cell, cell_text, current_row
                    if tag in ("td", "th"):
                        in_cell = True
                        cell_text = ""
                def handle_endtag(self, tag: str) -> None:
                    nonlocal in_cell, cell_text, current_row, html_rows
                    if tag in ("td", "th"):
                        in_cell = False
                        current_row.append(cell_text.strip())
                    elif tag == "tr" and current_row:
                        html_rows.append(tuple(current_row))
                        current_row = []
                def handle_data(self, data: str) -> None:
                    nonlocal cell_text
                    if in_cell:
                        cell_text += data

            parser = TableParser()
            parser.feed(text)
            if current_row:
                html_rows.append(tuple(current_row))
            if html_rows:
                logger.info(f"HTML success: {len(html_rows)} rows")
                return html_rows
    except Exception as exc:
        errors.append(f"html: {exc}")

    # Strategy 4: Try extracting CSV from inside a ZIP (some encrypted xlsx are just ZIPs)
    try:
        if contents[:2] == b'PK':
            zf = zipfile.ZipFile(io.BytesIO(contents))
            for name in zf.namelist():
                if name.lower().endswith(('.csv', '.txt')):
                    csv_bytes = zf.read(name)
                    text = csv_bytes.decode("utf-8-sig", errors="replace")
                    reader = csv.reader(io.StringIO(text))
                    rows = [tuple(row) for row in reader if any(cell.strip() for cell in row)]
                    if rows:
                        logger.info(f"ZIP-CSV success: {len(rows)} rows from {name}")
                        return rows
    except Exception as exc:
        errors.append(f"zip-csv: {exc}")

    error_detail = "; ".join(errors) if errors else "unknown format"
    raise ValueError(
        f"Cannot read the file ({len(contents)} bytes received). "
        f"Details: {error_detail}. "
        f"Supported formats: .xlsx, .xls, .csv. "
        f"If the file has IRM/DRM protection (Access workbook programmatically = No), "
        f"open it in Excel, Save As CSV (.csv), then upload the CSV file."
    )



# ---- Read endpoints ----

@router.get("/ngdc-datacenters")
async def list_ngdc_datacenters():
    """Return NGDC datacenters enriched with full neighbourhood objects for the UI.

    The raw JSON stores neighbourhoods as ID strings (e.g. ["NH01","NH02"]).
    The frontend NGDCSidebar expects objects: {name, subnets: string[]}.
    This endpoint resolves those IDs and also computes rule_count + ip_groups.
    """
    raw_dcs = await get_ngdc_datacenters()
    all_nhs = await get_neighbourhoods()
    all_rules = await get_rules()
    nh_lookup: dict[str, dict] = {nh["nh_id"]: nh for nh in all_nhs}

    enriched = []
    for dc in raw_dcs:
        dc_code = dc.get("code", "")
        # Resolve neighbourhood IDs to objects with subnets filtered to this DC
        nh_ids = dc.get("neighbourhoods", [])
        resolved_nhs = []
        ip_groups: list[dict] = []
        for nh_id in nh_ids:
            if isinstance(nh_id, str):
                nh_data = nh_lookup.get(nh_id)
                if nh_data:
                    # Filter IP ranges to those belonging to this DC
                    dc_ranges = [r for r in nh_data.get("ip_ranges", []) if r.get("dc") == dc_code]
                    subnets = [r["cidr"] for r in dc_ranges]
                    resolved_nhs.append({
                        "name": f"{nh_id} - {nh_data.get('name', '')}",
                        "nh_id": nh_id,
                        "security_zones": nh_data.get("security_zones", []),
                        "subnets": subnets,
                    })
                    # Also build ip_groups entries
                    if dc_ranges:
                        ip_groups.append({
                            "name": f"{nh_id} Groups",
                            "entries": [
                                {"name": r.get("description", r["cidr"]), "cidr": r["cidr"]}
                                for r in dc_ranges
                            ],
                        })
                else:
                    resolved_nhs.append({"name": nh_id, "nh_id": nh_id, "zone": "", "subnets": []})
            elif isinstance(nh_id, dict):
                # Already an object – ensure subnets key exists
                if "subnets" not in nh_id:
                    nh_id["subnets"] = []
                resolved_nhs.append(nh_id)

        # Count rules assigned to this DC
        rule_count = sum(1 for r in all_rules if r.get("datacenter") == dc_code)

        enriched.append({
            **dc,
            "neighbourhoods": resolved_nhs,
            "ip_groups": ip_groups,
            "rule_count": rule_count,
        })
    return enriched


@router.get("/security-zones")
async def list_security_zones():
    return await get_security_zones()


@router.get("/predefined-destinations")
async def list_predefined_destinations():
    return await get_predefined_destinations()


@router.get("/neighbourhoods")
async def list_neighbourhoods():
    return await get_neighbourhoods()


@router.get("/legacy-datacenters")
async def list_legacy_datacenters():
    return await get_legacy_datacenters()


@router.get("/applications")
async def list_applications():
    return await get_applications()


@router.get("/environments")
async def list_environments():
    return await get_environments()


@router.get("/chg-requests")
async def list_chg_requests():
    return await get_chg_requests()


@router.post("/chg-requests")
async def create_new_chg_request(data: dict):
    return await create_chg_request(data)


@router.get("/org-config")
async def get_org_config_endpoint():
    config = await get_org_config()
    if not config:
        raise HTTPException(status_code=404, detail="Org config not found")
    return config


@router.get("/policy-matrix")
async def list_policy_matrix():
    return await get_policy_matrix()


@router.get("/policy-matrix/heritage-dc")
async def list_heritage_dc_matrix():
    return await get_heritage_dc_matrix()


@router.get("/policy-matrix/ngdc-prod")
async def list_ngdc_prod_matrix():
    return await get_ngdc_prod_matrix()


@router.get("/policy-matrix/nonprod")
async def list_nonprod_matrix():
    return await get_nonprod_matrix()


@router.get("/policy-matrix/all")
async def list_all_policy_matrices():
    return {
        "heritage_dc": await get_heritage_dc_matrix(),
        "ngdc_prod": await get_ngdc_prod_matrix(),
        "nonprod": await get_nonprod_matrix(),
        "combined": await get_policy_matrix(),
    }


@router.get("/naming-standards")
async def list_naming_standards():
    return get_naming_standards_info()


# ---- Naming Standards utilities ----

@router.post("/naming-standards/validate")
async def validate_naming(data: dict):
    name = data.get("name", "")
    return validate_name(name)


@router.post("/naming-standards/generate")
async def generate_name(data: dict):
    name_type = data.get("type", "group")
    # Prefer app_distributed_id for {APP} placeholder; fall back to app_id
    app_id = data.get("app_distributed_id", "") or data.get("app_id", "")
    nh = data.get("nh", "NH01")
    sz = data.get("sz", "GEN")
    if name_type == "group":
        subtype = data.get("subtype", "APP")
        return generate_group_name(app_id, nh, sz, subtype)
    elif name_type == "server":
        server_name = data.get("server_name", "SRV01")
        return generate_server_name(app_id, nh, sz, server_name)
    elif name_type == "subnet":
        descriptor = data.get("descriptor", "NET")
        return generate_subnet_name(app_id, nh, sz, descriptor)
    return {"error": "Invalid type. Use 'group', 'server', or 'subnet'"}


@router.post("/naming-standards/suggest")
async def suggest_name(data: dict):
    legacy_name = data.get("legacy_name", "")
    # Prefer app_distributed_id for {APP} placeholder; fall back to app_id
    app_id = data.get("app_distributed_id", "") or data.get("app_id", "APP")
    nh = data.get("nh", "NH01")
    sz = data.get("sz", "GEN")
    return suggest_standard_name(legacy_name, app_id, nh, sz)


@router.post("/naming-standards/determine-zone")
async def determine_zone(data: dict):
    return determine_security_zone(
        paa_zone=data.get("paa_zone", False),
        exposure=data.get("exposure", "Internal"),
        pci_pan=data.get("pci_pan", False),
        pci_track_data=data.get("pci_track_data", False),
        pci_cvv_pin=data.get("pci_cvv_pin", False),
        deployment_type=data.get("deployment_type", "ON_PREMISE"),
        critical_payment=data.get("critical_payment", False),
        data_classification=data.get("data_classification", "INTERNAL"),
        criticality_rating=data.get("criticality_rating", 3),
        environment=data.get("environment", "Production"),
    )


# ---- CRUD: Neighbourhoods ----

@router.post("/neighbourhoods")
async def create_nh(data: dict):
    return await create_neighbourhood(data)


@router.put("/neighbourhoods/{nh_id}")
async def update_nh(nh_id: str, data: dict):
    result = await update_neighbourhood(nh_id, data)
    if not result:
        raise HTTPException(status_code=404, detail="Neighbourhood not found")
    return result


@router.delete("/neighbourhoods/{nh_id}")
async def delete_nh(nh_id: str):
    if not await delete_neighbourhood(nh_id):
        raise HTTPException(status_code=404, detail="Neighbourhood not found")
    return {"message": "Neighbourhood deleted"}


# ---- CRUD: Security Zones ----

@router.post("/security-zones")
async def create_sz(data: dict):
    return await create_security_zone(data)


@router.put("/security-zones/{code}")
async def update_sz(code: str, data: dict):
    result = await update_security_zone(code, data)
    if not result:
        raise HTTPException(status_code=404, detail="Security zone not found")
    return result


@router.delete("/security-zones/{code}")
async def delete_sz(code: str):
    if not await delete_security_zone(code):
        raise HTTPException(status_code=404, detail="Security zone not found")
    return {"message": "Security zone deleted"}


# ---- CRUD: Applications ----

@router.post("/applications")
async def create_app(data: dict):
    return await create_application(data)


@router.put("/applications/{app_id}")
async def update_app(app_id: str, data: dict):
    result = await update_application(app_id, data)
    if not result:
        raise HTTPException(status_code=404, detail="Application not found")
    return result


@router.delete("/applications/{app_id}")
async def delete_app(app_id: str):
    if not await delete_application(app_id):
        raise HTTPException(status_code=404, detail="Application not found")
    return {"message": "Application deleted"}


# ---- CRUD: Datacenters ----

@router.post("/ngdc-datacenters")
async def create_ngdc_dc(data: dict):
    return await create_datacenter(data, "ngdc")


@router.put("/ngdc-datacenters/{code}")
async def update_ngdc_dc(code: str, data: dict):
    result = await update_datacenter(code, data, "ngdc")
    if not result:
        raise HTTPException(status_code=404, detail="Datacenter not found")
    return result


@router.delete("/ngdc-datacenters/{code}")
async def delete_ngdc_dc(code: str):
    if not await delete_datacenter(code, "ngdc"):
        raise HTTPException(status_code=404, detail="Datacenter not found")
    return {"message": "Datacenter deleted"}


@router.post("/legacy-datacenters")
async def create_legacy_dc(data: dict):
    return await create_datacenter(data, "legacy")


@router.put("/legacy-datacenters/{code}")
async def update_legacy_dc(code: str, data: dict):
    result = await update_datacenter(code, data, "legacy")
    if not result:
        raise HTTPException(status_code=404, detail="Legacy datacenter not found")
    return result


@router.delete("/legacy-datacenters/{code}")
async def delete_legacy_dc(code: str):
    if not await delete_datacenter(code, "legacy"):
        raise HTTPException(status_code=404, detail="Legacy datacenter not found")
    return {"message": "Legacy datacenter deleted"}


# ---- CRUD: Predefined Destinations ----

@router.post("/predefined-destinations")
async def create_dest(data: dict):
    return await create_predefined_destination(data)


@router.put("/predefined-destinations/{name}")
async def update_dest(name: str, data: dict):
    result = await update_predefined_destination(name, data)
    if not result:
        raise HTTPException(status_code=404, detail="Destination not found")
    return result


@router.delete("/predefined-destinations/{name}")
async def delete_dest(name: str):
    if not await delete_predefined_destination(name):
        raise HTTPException(status_code=404, detail="Destination not found")
    return {"message": "Destination deleted"}


# ---- CRUD: Environments ----

@router.post("/environments")
async def create_env(data: dict):
    return await create_environment(data)


@router.delete("/environments/{code}")
async def delete_env(code: str):
    if not await delete_environment(code):
        raise HTTPException(status_code=404, detail="Environment not found")
    return {"message": "Environment deleted"}


# ---- CRUD: Org Config ----

@router.put("/org-config")
async def update_org(data: dict):
    result = await update_org_config(data)
    if not result:
        raise HTTPException(status_code=500, detail="Failed to update org config")
    return result


# ---- CRUD: Naming Standards ----

@router.put("/naming-standards")
async def update_standards(data: dict):
    result = await update_naming_standards(data)
    if not result:
        raise HTTPException(status_code=500, detail="Failed to update naming standards")
    return result


# ---- CRUD: Policy Matrix ----

@router.post("/policy-matrix")
async def create_policy(data: dict):
    return await create_policy_entry(data)


@router.delete("/policy-matrix/{source_zone}/{dest_zone}")
async def delete_policy(source_zone: str, dest_zone: str):
    if not await delete_policy_entry(source_zone, dest_zone):
        raise HTTPException(status_code=404, detail="Policy entry not found")
    return {"message": "Policy entry deleted"}


# ---- CRUD: Groups ----

@router.get("/groups")
async def list_groups(app_id: str | None = None):
    groups = await get_groups()
    if app_id:
        groups = [g for g in groups if g.get("app_id") == app_id or g.get("app_distributed_id") == app_id]
    return groups


@router.get("/groups/{name:path}")
async def get_group_endpoint(name: str):
    group = await get_group(name)
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    return group


@router.post("/groups")
async def create_new_group(data: dict):
    return await create_group(data)


@router.put("/groups/{name:path}")
async def update_existing_group(name: str, data: dict):
    result = await update_group(name, data)
    if not result:
        raise HTTPException(status_code=404, detail="Group not found")
    return result


@router.delete("/groups/{name:path}")
async def delete_existing_group(name: str):
    if not await delete_group(name):
        raise HTTPException(status_code=404, detail="Group not found")
    return {"message": "Group deleted"}


@router.post("/groups/{name:path}/members")
async def add_member_to_group(name: str, data: dict):
    result = await add_group_member(name, data)
    if not result:
        raise HTTPException(status_code=404, detail="Group not found")
    return result


@router.delete("/groups/{name:path}/members/{member_value}")
async def remove_member_from_group(name: str, member_value: str):
    result = await remove_group_member(name, member_value)
    if not result:
        raise HTTPException(status_code=404, detail="Group or member not found")
    return result


# ---- Legacy Rules (for Migration Studio & Firewall Management) ----

@router.get("/legacy-rules")
async def list_legacy_rules(app_id: str | None = None, exclude_migrated: bool = False, migration_only: bool = False):
    rules = await get_legacy_rules()
    if migration_only:
        # Show ALL imported legacy rules in the Migration module.
        # Previously this filtered for ngdc_imported==True, but rules imported
        # via Excel don't carry that flag — the user expects to see every
        # imported rule available for migration.
        pass
    if app_id:
        rules = [r for r in rules if str(r.get("app_id")) == str(app_id) or r.get("app_distributed_id") == app_id]
    if exclude_migrated:
        rules = [r for r in rules if r.get("migration_status") != "Completed"]
    return rules


@router.get("/legacy-rules/{rule_id}")
async def get_legacy_rule_detail(rule_id: str):
    rules = await get_legacy_rules()
    rule = next((r for r in rules if r["id"] == rule_id), None)
    if not rule:
        raise HTTPException(status_code=404, detail="Legacy rule not found")
    return rule


@router.put("/legacy-rules/{rule_id}")
async def update_legacy_rule_endpoint(rule_id: str, data: dict):
    result = await update_legacy_rule(rule_id, data)
    if not result:
        raise HTTPException(status_code=404, detail="Legacy rule not found")
    return result


@router.delete("/legacy-rules/clear-all")
async def clear_all_legacy_rules_endpoint():
    """Delete non-migrated legacy rules for a fresh re-import. Preserves migrated/in-progress rules and all mappings."""
    count = await clear_all_legacy_rules()
    return {"message": f"Deleted {count} non-migrated rules", "deleted": count}


@router.delete("/legacy-rules/{rule_id}")
async def delete_legacy_rule_endpoint(rule_id: str):
    ok = await delete_legacy_rule(rule_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Legacy rule not found")
    return {"message": "Deleted"}


def _normalize_header(h: str) -> str:
    """Normalize a header string for flexible matching: lowercase, strip whitespace, collapse spaces."""
    return " ".join(str(h).strip().lower().split())


def _strip_all(s: str) -> str:
    """Remove ALL non-alphanumeric characters for ultra-flexible matching."""
    return "".join(c for c in s.lower() if c.isalnum())


# Known header → internal field name mapping (normalised keys with spaces).
# _KNOWN_COL_MAP_NOSPACE has the same entries but with all spaces removed for
# matching concatenated headers like "Appid", "AppName", "ActionType", etc.
_KNOWN_COL_MAP: dict[str, str] = {
    # ── Standard rule fields (spaced form, e.g. "App ID", "Rule Source") ──
    "app id": "app_id",
    "app current distributed id": "app_distributed_id",
    "app distributed id": "app_distributed_id",
    "app name": "app_name",
    "inventory item": "inventory_item",
    "policy name": "policy_name",
    "rule global": "rule_global",
    "global rule": "rule_global",
    "rule action": "rule_action",
    "rule source": "rule_source",
    "rule source expanded": "rule_source_expanded",
    "rule source zone": "rule_source_zone",
    "rule destination": "rule_destination",
    "rule destination expanded": "rule_destination_expanded",
    "rule destination zone": "rule_destination_zone",
    "rule service": "rule_service",
    "rule service expanded": "rule_service_expanded",
    "rn": "rn",
    "rc": "rc",
    # ── Concatenated / PascalCase / alternate header names ──
    # App identifiers
    "appid": "app_id",
    "appname": "app_name",
    "appnme": "app_name",           # real variant seen in enterprise exports
    "app nme": "app_name",           # spaced variant
    "appdistributedid": "app_distributed_id",
    "appcurrentdistributedid": "app_distributed_id",
    "distributedid": "app_distributed_id",
    "distributed id": "app_distributed_id",
    "dist id": "app_distributed_id",
    "distid": "app_distributed_id",
    # App metadata (enterprise exports)
    "app asset status": "app_asset_status",
    "appassetstatus": "app_asset_status",
    "app portfolio": "app_portfolio",
    "appportfolio": "app_portfolio",
    "app manager": "app_manager",
    "appmanager": "app_manager",
    "appmana": "app_manager",        # real variant seen in enterprise exports
    "firewall name": "firewall_name",
    "firewallname": "firewall_name",
    # Inventory / policy
    "inventoryitem": "inventory_item",
    "policyname": "policy_name",
    "access policy": "policy_name",
    "accesspolicy": "policy_name",
    # Global rule
    "ruleglobal": "rule_global",
    "globalrule": "rule_global",
    "globalru": "rule_global",       # real variant seen in enterprise exports
    "global": "rule_global",
    # Action
    "ruleaction": "rule_action",
    "actiontype": "rule_action",
    "action type": "rule_action",
    "actiontyp": "rule_action",      # real variant seen in enterprise exports
    "action": "rule_action",
    # Source (semantic: Source = summary, SourceDetail/SourceExpanded = full list)
    "rulesource": "rule_source",
    "source": "rule_source",
    "src": "rule_source",
    "rulesourceexpanded": "rule_source_expanded",
    "sourceexpanded": "rule_source_expanded",
    "sourcedetail": "rule_source_expanded",
    "source detail": "rule_source_expanded",
    "source expanded": "rule_source_expanded",
    "srcdetail": "rule_source_expanded",
    "srcexpanded": "rule_source_expanded",
    "rulesourcezone": "rule_source_zone",
    "sourcezone": "rule_source_zone",
    "source zone": "rule_source_zone",
    "srczone": "rule_source_zone",
    # Destination (semantic: same pattern as source)
    "ruledestination": "rule_destination",
    "destination": "rule_destination",
    "dest": "rule_destination",
    "dst": "rule_destination",
    "ruledestinationexpanded": "rule_destination_expanded",
    "destinationexpanded": "rule_destination_expanded",
    "destinationdetail": "rule_destination_expanded",
    "destination detail": "rule_destination_expanded",
    "destination expanded": "rule_destination_expanded",
    "destdetail": "rule_destination_expanded",
    "destexpanded": "rule_destination_expanded",
    "dstdetail": "rule_destination_expanded",
    "ruledestinationzone": "rule_destination_zone",
    "destinationzone": "rule_destination_zone",
    "destination zone": "rule_destination_zone",
    "destzone": "rule_destination_zone",
    "dstzone": "rule_destination_zone",
    # Service (semantic: same pattern)
    "ruleservice": "rule_service",
    "service": "rule_service",
    "svc": "rule_service",
    "ruleserviceexpanded": "rule_service_expanded",
    "serviceexpanded": "rule_service_expanded",
    "servicedetail": "rule_service_expanded",
    "service detail": "rule_service_expanded",
    "service expanded": "rule_service_expanded",
    "svcdetail": "rule_service_expanded",
    "svcexpanded": "rule_service_expanded",
}

# Build a no-space lookup for ultra-flexible matching (catches "AppId", "App_ID", "APP-ID", etc.)
_KNOWN_COL_MAP_NOSPACE: dict[str, str] = {}
for _k, _v in _KNOWN_COL_MAP.items():
    ns = _strip_all(_k)
    if ns not in _KNOWN_COL_MAP_NOSPACE:
        _KNOWN_COL_MAP_NOSPACE[ns] = _v


def _header_to_field(header: str) -> str:
    """Convert an Excel header to an internal field name.
    Known headers map to their predefined field names.
    Tries: exact normalised match → no-space match → slugified fallback.
    Unknown headers are slugified: lowercased, spaces→underscores, non-alnum stripped."""
    norm = _normalize_header(header)
    if norm in _KNOWN_COL_MAP:
        return _KNOWN_COL_MAP[norm]
    # Try no-space match (catches "Appid", "AppName", "ActionType", etc.)
    ns = _strip_all(header)
    if ns in _KNOWN_COL_MAP_NOSPACE:
        return _KNOWN_COL_MAP_NOSPACE[ns]
    # Slugify: lowercase, replace spaces/special chars with underscore
    slug = norm.replace(" ", "_").replace("-", "_")
    slug = "".join(c for c in slug if c.isalnum() or c == "_")
    # Remove leading/trailing/double underscores
    while "__" in slug:
        slug = slug.replace("__", "_")
    return slug.strip("_") or f"col_{id(header)}"


@router.post("/legacy-rules/import")
async def import_legacy_rules_excel(file: UploadFile = File(...), environment: str = Form(default="")):
    """Import legacy rules from Excel (.xlsx/.xls) or CSV file with deduplication.
    ALL columns from the spreadsheet are preserved as-is.  Known columns
    (App ID, Rule Source, etc.) are mapped to canonical internal names;
    unknown columns are imported using a slugified version of the header.
    Accepts optional 'environment' form field from the dropdown selection."""
    import logging
    logger = logging.getLogger("excel_import")
    fname = file.filename or ""
    if not fname.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="Supported formats: .xlsx, .xls, .csv")
    contents = await file.read()
    logger.info(f"Import request: file={fname}, size={len(contents)} bytes")
    try:
        rows = _parse_excel_bytes(contents, filename=fname)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty workbook")
    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    # Build field name list — every column is imported
    field_names: list[str] = []
    mapped_headers: list[str] = []
    unmapped_headers: list[str] = []
    for h in raw_headers:
        fn = _header_to_field(h)
        field_names.append(fn)
        if _normalize_header(h) in _KNOWN_COL_MAP or _strip_all(h) in _KNOWN_COL_MAP_NOSPACE:
            mapped_headers.append(h)
        elif h:
            unmapped_headers.append(h)
    logger.info(f"Headers ({len(raw_headers)}): mapped={len(mapped_headers)}, unmapped={len(unmapped_headers)}, total_data_rows={len(rows) - 1}")
    parsed_rules: list[dict] = []
    skipped_empty = 0
    for row_vals in rows[1:]:
        if not any(v is not None and str(v).strip() for v in row_vals):
            skipped_empty += 1
            continue
        rule: dict = {}
        for i, fn in enumerate(field_names):
            if i < len(row_vals) and fn:
                val = row_vals[i]
                rule[fn] = str(val).strip() if val is not None else ""
        rule["is_standard"] = False
        rule.setdefault("migration_status", "Not Started")
        # Set environment from dropdown selection (passed as form field)
        if environment:
            rule["environment"] = environment
        parsed_rules.append(rule)
    logger.info(f"Parsed {len(parsed_rules)} rules (skipped {skipped_empty} empty rows)")
    result = await import_legacy_rules(parsed_rules)
    # Add diagnostics to the response
    result["parsed_rows"] = len(parsed_rules)
    result["skipped_empty"] = skipped_empty
    result["total_file_rows"] = len(rows) - 1
    result["headers_found"] = raw_headers
    result["mapped_headers"] = mapped_headers
    result["unmapped_headers"] = unmapped_headers
    return result


@router.post("/legacy-rules/import-json")
async def import_legacy_rules_json(file: UploadFile = File(...)):
    """Import legacy rules from a JSON file with deduplication.
    Expects a JSON array of objects with same field names as Excel columns
    (either Excel header names or internal field names are accepted)."""
    if not file.filename or not file.filename.endswith(".json"):
        raise HTTPException(status_code=400, detail="Only .json files are supported")
    contents = await file.read()
    try:
        data = json_lib.loads(contents)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid JSON: {exc}")
    if not isinstance(data, list):
        raise HTTPException(status_code=400, detail="JSON must be an array of rule objects")
    if not data:
        raise HTTPException(status_code=400, detail="Empty JSON array")

    # Accept both Excel header names and internal field names
    col_map = {
        "App ID": "app_id", "app_id": "app_id",
        "App Current Distributed ID": "app_distributed_id", "app_distributed_id": "app_distributed_id",
        "App Name": "app_name", "app_name": "app_name",
        "Inventory Item": "inventory_item", "inventory_item": "inventory_item",
        "Policy Name": "policy_name", "policy_name": "policy_name",
        "Rule Global": "rule_global", "rule_global": "rule_global",
        "Rule Action": "rule_action", "rule_action": "rule_action",
        "Rule Source": "rule_source", "rule_source": "rule_source",
        "Rule Source Expanded": "rule_source_expanded", "rule_source_expanded": "rule_source_expanded",
        "Rule Source Zone": "rule_source_zone", "rule_source_zone": "rule_source_zone",
        "Rule Destination": "rule_destination", "rule_destination": "rule_destination",
        "Rule Destination Expanded": "rule_destination_expanded", "rule_destination_expanded": "rule_destination_expanded",
        "Rule Destination Zone": "rule_destination_zone", "rule_destination_zone": "rule_destination_zone",
        "Rule Service": "rule_service", "rule_service": "rule_service",
        "Rule Service Expanded": "rule_service_expanded", "rule_service_expanded": "rule_service_expanded",
        "RN": "rn", "rn": "rn",
        "RC": "rc", "rc": "rc",
    }
    parsed_rules = []
    for item in data:
        if not isinstance(item, dict):
            continue
        rule: dict = {}
        for key, val in item.items():
            mapped = col_map.get(key)
            if mapped:
                rule[mapped] = str(val) if val is not None else ""
        rule["is_standard"] = False
        rule["migration_status"] = "Not Started"
        parsed_rules.append(rule)
    result = await import_legacy_rules(parsed_rules)
    return result


@router.get("/legacy-rules/export-excel")
async def export_legacy_rules_excel(app_id: str = ""):
    """Export imported legacy rules as an Excel (.xlsx) file."""
    rules = await get_legacy_rules()
    if app_id:
        rules = [r for r in rules if str(r.get("app_id", "")) == str(app_id) or r.get("app_distributed_id", "") == app_id]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Legacy Rules"
    headers = [
        "App ID", "App Current Distributed ID", "App Name", "Inventory Item",
        "Policy Name", "Rule Global", "Rule Action",
        "Rule Source", "Rule Source Expanded", "Rule Source Zone",
        "Rule Destination", "Rule Destination Expanded", "Rule Destination Zone",
        "Rule Service", "Rule Service Expanded", "RN", "RC", "Migration Status",
    ]
    field_map = [
        "app_id", "app_distributed_id", "app_name", "inventory_item",
        "policy_name", "rule_global", "rule_action",
        "rule_source", "rule_source_expanded", "rule_source_zone",
        "rule_destination", "rule_destination_expanded", "rule_destination_zone",
        "rule_service", "rule_service_expanded", "rn", "rc", "migration_status",
    ]
    ws.append(headers)
    for rule in rules:
        ws.append([rule.get(f, "") for f in field_map])

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    filename = f"legacy_rules_{app_id}.xlsx" if app_id else "legacy_rules.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/legacy-rules/imported-apps")
async def get_imported_apps():
    """Return unique apps from imported legacy rules with their mapping status.
    For each app, indicates whether it already has app-dc-mappings (NH/SZ/DC)
    and lists its existing component mappings if any.
    Reads from BOTH seed and live directories so apps are always visible regardless of data mode."""
    from app.database import get_app_dc_mappings, get_all_legacy_rules_across_modes
    rules = await get_all_legacy_rules_across_modes()
    app_dc_mappings = await get_app_dc_mappings()

    # Build lookup of existing mappings by app_id
    mapped: dict[str, list[dict]] = {}
    for m in app_dc_mappings:
        aid = str(m.get("app_id", ""))
        mapped.setdefault(aid, []).append(m)

    # Extract unique apps from legacy rules
    seen: set[str] = set()
    apps: list[dict] = []
    for r in rules:
        app_id = str(r.get("app_id", "")).strip()
        if not app_id or app_id in seen:
            continue
        seen.add(app_id)
        existing_mappings = mapped.get(app_id, [])
        apps.append({
            "app_id": app_id,
            "app_name": r.get("app_name", ""),
            "app_distributed_id": r.get("app_distributed_id", ""),
            "rule_count": sum(1 for lr in rules if str(lr.get("app_id", "")).strip() == app_id),
            "has_mapping": len(existing_mappings) > 0,
            "components": existing_mappings,
        })
    return apps


@router.post("/legacy-rules/migrate")
async def migrate_rules_to_ngdc(data: dict):
    """Migrate selected rules to NGDC standards."""
    rule_ids = data.get("rule_ids", [])
    if not rule_ids:
        raise HTTPException(status_code=400, detail="rule_ids required")
    migrated = []
    for rid in rule_ids:
        result = await migrate_rule_to_ngdc(rid)
        if result:
            migrated.append(result)
    return {"migrated": len(migrated), "rules": migrated}


@router.post("/legacy-rules/submit-for-review")
async def submit_legacy_rules_for_review(data: dict):
    """Submit selected legacy rules for migration review."""
    rule_ids = data.get("rule_ids", [])
    comments = data.get("comments", "")
    if not rule_ids:
        raise HTTPException(status_code=400, detail="rule_ids required")
    reviews = await create_migration_review(rule_ids, comments)
    return {"submitted": len(reviews), "reviews": reviews}


@router.get("/migration-history")
async def list_migration_history():
    return await get_migration_history()


@router.get("/legacy-rules/migrated")
async def list_migrated_rules():
    """Return only rules that have been migrated to NGDC (for Firewall Studio)."""
    rules = await get_legacy_rules()
    return [r for r in rules if r.get("migration_status") == "Completed"]


# ---- Rule Modification with Delta Tracking ----

@router.post("/legacy-rules/{rule_id}/modify")
async def modify_legacy_rule(rule_id: str, data: dict):
    """Create a rule modification request with delta tracking."""
    modifications = data.get("modifications", {})
    comments = data.get("comments", "")
    result = await create_rule_modification(rule_id, modifications, comments)
    if not result:
        raise HTTPException(status_code=404, detail="Legacy rule not found")
    return result


@router.get("/rule-modifications")
async def list_rule_modifications(rule_id: str | None = None):
    return await get_rule_modifications(rule_id)


@router.post("/rule-modifications/{mod_id}/approve")
async def approve_mod(mod_id: str, data: dict):
    notes = data.get("notes", "")
    result = await approve_rule_modification(mod_id, notes)
    if not result:
        raise HTTPException(status_code=404, detail="Modification not found")
    return result


@router.post("/rule-modifications/{mod_id}/reject")
async def reject_mod(mod_id: str, data: dict):
    notes = data.get("notes", "")
    if not notes:
        raise HTTPException(status_code=400, detail="Rejection notes required")
    result = await reject_rule_modification(mod_id, notes)
    if not result:
        raise HTTPException(status_code=404, detail="Modification not found")
    return result


# ---- Legacy Rule Compiler ----

@router.post("/legacy-rules/{rule_id}/compile")
async def compile_legacy(rule_id: str, vendor: str = "generic"):
    result = await compile_legacy_rule(rule_id, vendor)
    if not result:
        raise HTTPException(status_code=404, detail="Legacy rule not found")
    return result


# ---- NGDC Recommendations ----

@router.get("/legacy-rules/{rule_id}/ngdc-recommendations")
async def get_recommendations(rule_id: str):
    result = await get_ngdc_recommendations(rule_id)
    if not result:
        raise HTTPException(status_code=404, detail="Legacy rule not found")
    return result


# ---- Birthright Validation ----

@router.post("/birthright/validate")
async def validate_birthright_rule(data: dict):
    return await validate_birthright(data)


@router.get("/birthright/matrix")
async def get_birthright_matrices():
    return await get_birthright_matrix()


@router.put("/birthright/matrix/{matrix_type}")
async def update_matrix(matrix_type: str, data: dict):
    entries = data.get("entries", [])
    result = await update_birthright_matrix(matrix_type, entries)
    if not result and result != []:
        raise HTTPException(status_code=400, detail="Invalid matrix type")
    return result


@router.post("/birthright/matrix/{matrix_type}")
async def add_matrix_entry(matrix_type: str, data: dict):
    return await add_birthright_entry(matrix_type, data)


@router.delete("/birthright/matrix/{matrix_type}/{index}")
async def delete_matrix_entry(matrix_type: str, index: int):
    if not await delete_birthright_entry(matrix_type, index):
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Entry deleted"}


# ---- Legacy-to-NGDC Organization Mappings ----

@router.get("/ngdc-mappings")
async def list_ngdc_mappings():
    from app.database import get_ngdc_mappings
    return await get_ngdc_mappings()


@router.post("/ngdc-mappings")
async def create_ngdc_mapping(data: dict):
    from app.database import add_ngdc_mapping
    return await add_ngdc_mapping(data)


@router.put("/ngdc-mappings/{mapping_id}")
async def update_ngdc_mapping_endpoint(mapping_id: str, data: dict):
    from app.database import update_ngdc_mapping
    result = await update_ngdc_mapping(mapping_id, data)
    if not result:
        raise HTTPException(status_code=404, detail="Mapping not found")
    return result


@router.delete("/ngdc-mappings/{mapping_id}")
async def delete_ngdc_mapping_endpoint(mapping_id: str):
    from app.database import delete_ngdc_mapping
    ok = await delete_ngdc_mapping(mapping_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Mapping not found")
    return {"message": "Mapping deleted"}


@router.post("/ngdc-mappings/import")
async def import_ngdc_mappings_excel(file: UploadFile = File(...)):
    """Import legacy-to-NGDC mappings from Excel or CSV."""
    from app.database import import_ngdc_mappings
    fname = file.filename or ""
    if not fname.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="Supported formats: .xlsx, .xls, .csv")
    contents = await file.read()
    try:
        rows = _parse_excel_bytes(contents, filename=fname)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty workbook")
    headers = list(rows[0])
    mappings = []
    for row_vals in rows[1:]:
        if not any(row_vals):
            continue
        m: dict = {}
        for i, h in enumerate(headers):
            if i < len(row_vals) and h:
                m[str(h).lower().replace(" ", "_")] = str(row_vals[i]) if row_vals[i] is not None else ""
        mappings.append(m)
    result = await import_ngdc_mappings(mappings)
    return result


@router.post("/ngdc-mappings/bulk")
async def bulk_save_ngdc_mappings(data: dict):
    """Save all mappings at once (from org admin)."""
    from app.database import save_ngdc_mappings
    mappings = data.get("mappings", [])
    return await save_ngdc_mappings(mappings)


# ---- Group Provisioning to Firewall Device ----

@router.post("/groups/provision/{app_id}")
async def provision_groups(app_id: str, data: dict):
    """Provision groups for an app to the firewall device (NGDC standards enforced)."""
    from app.database import provision_groups_to_device
    device_type = data.get("device_type", "palo_alto")
    return await provision_groups_to_device(app_id, device_type)


@router.get("/provisioning-history")
async def list_provisioning_history(app_id: str | None = None):
    from app.database import get_provisioning_history
    return await get_provisioning_history(app_id)


# ---- Enhanced Compile with Group Expansion ----

@router.post("/rules/{rule_id}/compile-expanded")
async def compile_rule_expanded(rule_id: str, data: dict):
    from app.database import compile_rule_with_expansion
    vendor = data.get("vendor", "generic")
    expand = data.get("expand_groups", True)
    result = await compile_rule_with_expansion(rule_id, vendor, expand)
    if not result:
        raise HTTPException(status_code=404, detail="Rule not found")
    return result


# ---- App-to-DC/NH/SZ Mapping (Organization Reference) ----

@router.get("/app-dc-mappings")
async def list_app_dc_mappings():
    from app.database import get_app_dc_mappings
    return await get_app_dc_mappings()


@router.get("/app-dc-mappings/{app_id}")
async def get_app_dc_mapping(app_id: str):
    from app.database import get_app_dc_mapping_by_app
    result = await get_app_dc_mapping_by_app(app_id)
    if not result:
        raise HTTPException(status_code=404, detail="App DC mapping not found")
    return result


@router.post("/app-dc-mappings")
async def create_app_dc_mapping(data: dict):
    from app.database import add_app_dc_mapping
    return await add_app_dc_mapping(data)


@router.put("/app-dc-mappings/{mapping_id}")
async def update_app_dc_mapping_endpoint(mapping_id: str, data: dict):
    from app.database import update_app_dc_mapping
    result = await update_app_dc_mapping(mapping_id, data)
    if not result:
        raise HTTPException(status_code=404, detail="Mapping not found")
    return result


@router.delete("/app-dc-mappings/{mapping_id}")
async def delete_app_dc_mapping_endpoint(mapping_id: str):
    from app.database import delete_app_dc_mapping
    ok = await delete_app_dc_mapping(mapping_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Mapping not found")
    return {"message": "Mapping deleted"}


@router.post("/app-dc-mappings/import")
async def import_app_dc_mappings_excel(file: UploadFile = File(...)):
    """Import App-to-DC/NH/SZ mappings from Excel or CSV."""
    from app.database import import_app_dc_mappings
    fname = file.filename or ""
    if not fname.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="Supported formats: .xlsx, .xls, .csv")
    contents = await file.read()
    try:
        rows = _parse_excel_bytes(contents, filename=fname)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty workbook")
    headers = list(rows[0])
    mappings = []
    for row_vals in rows[1:]:
        if not any(row_vals):
            continue
        m: dict = {}
        for i, h in enumerate(headers):
            if i < len(row_vals) and h:
                key = str(h).lower().replace(" ", "_")
                m[key] = str(row_vals[i]) if row_vals[i] is not None else ""
        mappings.append(m)
    result = await import_app_dc_mappings(mappings)
    return result


@router.post("/app-dc-mappings/bulk")
async def bulk_save_app_dc_mappings(data: dict):
    from app.database import save_app_dc_mappings
    mappings = data.get("mappings", [])
    return await save_app_dc_mappings(mappings)


# ---- NGDC Compliance Check ----

@router.post("/legacy-rules/check-compliance")
async def check_compliance(data: dict):
    """Check NGDC compliance for a single rule or batch of rules."""
    from app.database import check_ngdc_compliance, get_legacy_rules
    rule_ids = data.get("rule_ids", [])
    if not rule_ids:
        raise HTTPException(status_code=400, detail="rule_ids required")
    rules = await get_legacy_rules()
    results = []
    for rid in rule_ids:
        rule = next((r for r in rules if r["id"] == rid), None)
        if rule:
            result = await check_ngdc_compliance(rule)
            results.append(result)
    return results


# ---- Duplicate Detection ----

@router.post("/check-duplicates")
async def check_dups(data: dict):
    """Check for duplicate rules based on source+destination+service."""
    from app.database import check_duplicates
    source = data.get("source", "")
    destination = data.get("destination", "")
    service = data.get("service", "")
    exclude_id = data.get("exclude_id", "")
    duplicates = await check_duplicates(source, destination, service, exclude_id)
    return {"duplicates": duplicates, "count": len(duplicates)}


# ---- Import Rules to NGDC Standardization (from Network Firewall Request) ----

@router.post("/legacy-rules/import-to-ngdc")
async def import_to_ngdc(data: dict):
    """Import rules from Network Firewall Request to NGDC Standardization by app IDs."""
    from app.database import import_rules_to_ngdc_standardization
    app_ids = data.get("app_ids", [])
    if not app_ids:
        raise HTTPException(status_code=400, detail="app_ids required")
    return await import_rules_to_ngdc_standardization(app_ids)


# ---- Auto-Import Compliant Rules to Firewall Studio ----

@router.post("/legacy-rules/auto-import-to-studio")
async def auto_import_to_studio():
    """Auto-import NGDC-compliant rules from Network Firewall Request into Firewall Studio."""
    from app.database import auto_import_compliant_rules_to_studio
    return await auto_import_compliant_rules_to_studio()


# ---- Expand Groups in Rule ----

@router.get("/legacy-rules/{rule_id}/expanded")
async def get_expanded_rule(rule_id: str):
    """Get a legacy rule with groups expanded to show all IPs/ranges."""
    from app.database import expand_groups_in_rule, get_legacy_rules
    rules = await get_legacy_rules()
    rule = next((r for r in rules if r["id"] == rule_id), None)
    if not rule:
        raise HTTPException(status_code=404, detail="Legacy rule not found")
    return await expand_groups_in_rule(rule)


# ---- Create Migration Group ----

@router.post("/migration-groups")
async def create_mig_group(data: dict):
    """Create a new group during migration with NGDC naming standards."""
    from app.database import create_migration_group
    name = data.get("name", "")
    app_id = data.get("app_id", "")
    members = data.get("members", [])
    nh = data.get("nh", "")
    sz = data.get("sz", "")
    if not name:
        raise HTTPException(status_code=400, detail="Group name required")
    return await create_migration_group(name, app_id, members, nh, sz)


# ---- IP Mappings (Legacy DC <-> NGDC one-to-one) ----

@router.get("/ip-mappings")
async def list_ip_mappings(legacy_dc: str | None = None, app_id: str | None = None):
    from app.database import get_ip_mappings
    return await get_ip_mappings(legacy_dc, app_id)


@router.post("/ip-mappings")
async def create_ip_mapping(data: dict):
    from app.database import add_ip_mapping
    return await add_ip_mapping(data)


@router.put("/ip-mappings/{mapping_id}")
async def update_ip_mapping_endpoint(mapping_id: str, data: dict):
    from app.database import update_ip_mapping
    result = await update_ip_mapping(mapping_id, data)
    if not result:
        raise HTTPException(status_code=404, detail="IP mapping not found")
    return result


@router.delete("/ip-mappings/{mapping_id}")
async def delete_ip_mapping_endpoint(mapping_id: str):
    from app.database import delete_ip_mapping
    if not await delete_ip_mapping(mapping_id):
        raise HTTPException(status_code=404, detail="IP mapping not found")
    return {"message": "IP mapping deleted"}


@router.post("/ip-mappings/lookup")
async def lookup_ip_mapping(data: dict):
    """Look up the NGDC equivalent for a legacy IP address."""
    from app.database import lookup_ngdc_ip
    legacy_ip = data.get("legacy_ip", "")
    legacy_dc = data.get("legacy_dc", "")
    result = await lookup_ngdc_ip(legacy_ip, legacy_dc)
    if not result:
        return {"found": False, "message": f"No NGDC mapping found for {legacy_ip}"}
    return {"found": True, "mapping": result}


# ---- Firewall Boundary Analysis ----

@router.get("/firewall-boundaries")
async def get_firewall_boundaries(
    src_nh: str = "", src_sz: str = "", dst_nh: str = "", dst_sz: str = ""
):
    """Determine how many firewall boundaries a rule must cross based on
    the Logical Data Flow rules (NH/SZ placement)."""
    from app.database import determine_firewall_boundaries
    return await determine_firewall_boundaries(src_nh, src_sz, dst_nh, dst_sz)


@router.get("/logical-flow-rules")
async def get_logical_flow_rules():
    """Return the Logical Data Flow rule definitions."""
    from app.seed_data import LOGICAL_FLOW_RULES, SEGMENTED_ZONES
    return {"rules": LOGICAL_FLOW_RULES, "segmented_zones": list(SEGMENTED_ZONES)}


# ---- Egress/Ingress Compilation ----

@router.post("/compile/egress-ingress/{rule_id}")
async def compile_egress_ingress_endpoint(rule_id: str, vendor: str = "generic"):
    """Compile separate egress and ingress rules for cross-SZ combinations.
    Uses Logical Data Flow boundary analysis for device-specific compilation."""
    from app.database import compile_egress_ingress
    result = await compile_egress_ingress(rule_id, vendor)
    if not result:
        raise HTTPException(status_code=404, detail="Rule not found")
    return result


# ---- Resolved Policy Matrix ----

@router.get("/policy-matrix/resolved")
async def get_resolved_matrix(
    src_dc: str = "", src_nh: str = "", src_sz: str = "",
    dst_dc: str = "", dst_nh: str = "", dst_sz: str = "",
    environment: str = "Production",
):
    """Return policy matrix with Same/Any/Different resolved to real NH/SZ names."""
    from app.database import get_resolved_policy_matrix
    return await get_resolved_policy_matrix(src_dc, src_nh, src_sz, dst_dc, dst_nh, dst_sz, environment)


@router.get("/policy-matrix/preprod")
async def list_preprod_matrix():
    from app.database import get_preprod_matrix
    return await get_preprod_matrix()


# ---- Firewall Devices ----

@router.get("/firewall-devices")
async def list_firewall_devices():
    from app.database import get_firewall_devices
    return await get_firewall_devices()


@router.get("/firewall-devices/{device_id}")
async def get_firewall_device_endpoint(device_id: str):
    from app.database import get_firewall_device
    device = await get_firewall_device(device_id)
    if not device:
        raise HTTPException(status_code=404, detail="Firewall device not found")
    return device


@router.post("/firewall-devices")
async def create_firewall_device_endpoint(data: dict):
    from app.database import create_firewall_device
    return await create_firewall_device(data)


@router.put("/firewall-devices/{device_id}")
async def update_firewall_device_endpoint(device_id: str, data: dict):
    from app.database import update_firewall_device
    result = await update_firewall_device(device_id, data)
    if not result:
        raise HTTPException(status_code=404, detail="Firewall device not found")
    return result


@router.delete("/firewall-devices/{device_id}")
async def delete_firewall_device_endpoint(device_id: str):
    from app.database import delete_firewall_device
    ok = await delete_firewall_device(device_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Firewall device not found")
    return {"message": "Deleted"}


# ---- IP Mappings Import ----

@router.post("/ip-mappings/import")
async def import_ip_mappings_endpoint(data: dict):
    """Import multiple IP mappings at once, optionally for a specific app."""
    from app.database import import_ip_mappings
    records = data.get("mappings", [])
    app_id = data.get("app_id")
    return import_ip_mappings(records, app_id)


# ---- Firewall Device Patterns ----

@router.get("/firewall-device-patterns")
async def list_fw_device_patterns():
    """Return the generic firewall device naming patterns and DC vendor map."""
    from app.database import get_firewall_device_patterns, get_dc_vendor_map
    return {
        "patterns": await get_firewall_device_patterns(),
        "dc_vendor_map": await get_dc_vendor_map(),
    }


# ---- Separate JSON Storage (user-data/) — Migration Data & Studio Rules ----

@router.get("/user-data/summary")
async def get_user_data_summary():
    """Return summary of all separate user-data JSON files."""
    from app.database import get_all_user_data_files
    return await get_all_user_data_files()


@router.get("/user-data/migration")
async def get_migration_data_endpoint():
    """Get all migration data from separate JSON."""
    from app.database import get_migration_data
    return await get_migration_data()


@router.delete("/user-data/migration")
async def clear_migration_data_endpoint():
    """Clear all migration data from separate JSON for safe cleanup."""
    from app.database import clear_migration_data
    counts = await clear_migration_data()
    return {"status": "cleared", "cleared_counts": counts}


@router.get("/user-data/studio-rules")
async def get_studio_rules_endpoint():
    """Get all studio rules from separate JSON."""
    from app.database import get_studio_rules
    return await get_studio_rules()


@router.delete("/user-data/studio-rules")
async def clear_studio_rules_endpoint():
    """Clear all studio rules from separate JSON for safe cleanup."""
    from app.database import clear_studio_rules
    count = await clear_studio_rules()
    return {"status": "cleared", "count": count}


@router.delete("/user-data/studio-rules/{rule_id}")
async def delete_studio_rule_endpoint(rule_id: str):
    """Delete a specific studio rule from separate JSON."""
    from app.database import delete_studio_rule
    deleted = await delete_studio_rule(rule_id)
    if not deleted:
        raise HTTPException(404, "Studio rule not found")
    return {"status": "deleted", "rule_id": rule_id}


# ---- Hide Seed Data Toggle ----

@router.get("/hide-seed")
async def get_hide_seed_endpoint():
    """Return the current hide-seed-data setting."""
    from app.database import get_hide_seed
    return {"hide_seed": get_hide_seed()}


@router.post("/hide-seed")
async def set_hide_seed_endpoint(data: dict):
    """Toggle hide-seed-data. When True, only real/imported data is returned."""
    from app.database import set_hide_seed
    hide = bool(data.get("hide", False))
    result = set_hide_seed(hide)
    return {"hide_seed": result}


@router.get("/rules/real")
async def get_real_rules_only():
    """Return only real/user-created rules (from studio_rules.json + legacy imported).
    Use this endpoint when hide-seed-data is enabled."""
    from app.database import get_studio_rules, get_legacy_rules
    studio = await get_studio_rules()
    legacy = await get_legacy_rules()
    # Combine: studio rules are the primary real rules, legacy are imported
    studio_ids = {r.get("rule_id") for r in studio}
    combined = list(studio)
    for lr in legacy:
        if lr.get("rule_id") and lr["rule_id"] not in studio_ids:
            combined.append(lr)
    return combined


@router.get("/groups/real")
async def get_real_groups_only():
    """Return only groups that were created by the user (not seed data).
    Groups with _user_created=True or created after seed are considered real."""
    from app.database import get_groups
    groups = await get_groups()
    # Filter: only groups with _user_created flag or groups created via Studio
    real_groups = [g for g in groups if g.get("_user_created", False) or g.get("source") == "studio"]
    return real_groups


@router.get("/reviews/real")
async def get_real_reviews_only():
    """Return only reviews for real/user-created rules."""
    from app.database import get_reviews, get_studio_rules
    reviews = await get_reviews()
    studio = await get_studio_rules()
    studio_ids = {r.get("rule_id") for r in studio}
    real_reviews = [r for r in reviews if r.get("rule_id") in studio_ids]
    return real_reviews


# ---- Cleanup Endpoints (individual + one-click reset) ----

@router.delete("/user-data/all")
async def clear_all_user_data_endpoint():
    """One-click reset: clear ALL user/imported data across all stores.
    Seed reference data (NHs, SZs, policy matrix, etc.) is NOT affected."""
    from app.database import clear_all_user_data
    counts = await clear_all_user_data()
    return {"status": "cleared", "counts": counts}


@router.delete("/user-data/reviews")
async def clear_reviews_endpoint():
    """Clear all reviews."""
    from app.database import clear_reviews
    count = await clear_reviews()
    return {"status": "cleared", "count": count}


@router.delete("/user-data/groups")
async def clear_groups_endpoint():
    """Clear all groups."""
    from app.database import clear_user_groups
    count = await clear_user_groups()
    return {"status": "cleared", "count": count}


@router.delete("/user-data/firewall-rules")
async def clear_firewall_rules_endpoint():
    """Clear all firewall rules (from firewall_rules.json)."""
    from app.database import clear_firewall_rules
    count = await clear_firewall_rules()
    return {"status": "cleared", "count": count}


@router.delete("/user-data/modifications")
async def clear_modifications_endpoint():
    """Clear all rule modifications."""
    from app.database import clear_modifications
    count = await clear_modifications()
    return {"status": "cleared", "count": count}


@router.delete("/user-data/legacy-rules")
async def clear_legacy_rules_force_endpoint():
    """Force-clear ALL legacy rules (including migrated). For full reset."""
    from app.database import clear_all_legacy_rules_force
    count = await clear_all_legacy_rules_force()
    return {"status": "cleared", "count": count}


@router.delete("/user-data/by-app/{app_id}")
async def clear_data_by_app_endpoint(app_id: str):
    """Clear all data for a specific application across all stores."""
    from app.database import clear_data_by_app
    counts = await clear_data_by_app(app_id)
    return {"status": "cleared", "app_id": app_id, "counts": counts}


@router.delete("/user-data/by-env/{environment}")
async def clear_data_by_env_endpoint(environment: str):
    """Clear all data for a specific environment across all stores."""
    from app.database import clear_data_by_environment
    counts = await clear_data_by_environment(environment)
    return {"status": "cleared", "environment": environment, "counts": counts}


@router.get("/user-data/summary/by-app")
async def get_data_summary_by_app_endpoint():
    """Return per-app record counts for Data Management overview."""
    from app.database import get_data_summary_by_app
    return await get_data_summary_by_app()


@router.get("/user-data/summary/by-env")
async def get_data_summary_by_env_endpoint():
    """Return per-environment record counts for Data Management overview."""
    from app.database import get_data_summary_by_env
    return await get_data_summary_by_env()


# ---- Auto-populate NH/SZ/DC by Environment + App ----

@router.get("/filtered-nh-sz-dc")
async def get_filtered_nh_sz_dc_endpoint(environment: str, app_id: str | None = None):
    """Return NHs/SZs/DCs filtered by environment and optionally by app.
    Non-prod/preprod NHs are common for all apps; prod NHs are app-specific."""
    from app.database import get_filtered_nh_sz_dc
    return await get_filtered_nh_sz_dc(environment, app_id)


# ---- App Management Clear / Delta Import ----

@router.post("/applications/clear")
async def clear_app_management_endpoint():
    """Clear all imported (non-seed) application data."""
    from app.database import clear_app_management
    return await clear_app_management()


@router.post("/applications/import")
async def import_app_management_endpoint(file: UploadFile = File(...)):
    """Delta-based import of app management data from Excel/CSV.
    Uses app_distributed_id as the dedup key."""
    from app.database import import_app_management
    contents = await file.read()
    rows = _parse_excel_bytes(contents, file.filename or "")
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="File has no data rows")

    headers = [str(h).strip().lower().replace(" ", "_") for h in rows[0]]
    # Map common column name variations
    col_map = {
        "app_id": "app_id", "appid": "app_id", "app id": "app_id",
        "app_distributed_id": "app_distributed_id", "distributed_id": "app_distributed_id",
        "app distributed id": "app_distributed_id", "@work_id": "app_distributed_id",
        "@work id": "app_distributed_id", "work_id": "app_distributed_id",
        "app_name": "name", "name": "name", "application_name": "name", "app name": "name",
        "neighborhoods": "neighborhoods", "neighbourhood": "neighborhoods", "nhs": "neighborhoods", "nh": "neighborhoods",
        "szs": "szs", "sz": "szs", "security_zones": "szs", "security zones": "szs",
        "dcs": "dcs", "dc": "dcs", "data_centers": "dcs", "data centers": "dcs", "datacenters": "dcs",
        "snow_sysid": "snow_sysid", "sysid": "snow_sysid", "snow sysid": "snow_sysid",
        "servicenow_sysid": "snow_sysid",
        "owner": "owner", "criticality": "criticality", "pci_scope": "pci_scope",
    }

    mapped_headers = []
    for h in headers:
        mapped = col_map.get(h, h)
        mapped_headers.append(mapped)

    records = []
    for row in rows[1:]:
        rec: dict = {}
        for i, val in enumerate(row):
            if i < len(mapped_headers):
                rec[mapped_headers[i]] = str(val).strip() if val is not None else ""
        if rec.get("app_distributed_id"):
            records.append(rec)

    if not records:
        raise HTTPException(status_code=400, detail="No valid records found (app_distributed_id column required)")

    result = await import_app_management(records)
    return result


# ---- Standalone JSON Seed Export ----

@router.get("/export/seed-json")
async def export_seed_json():
    """Export all seed reference data as standalone JSON objects suitable for
    external consumption.  Returns neighbourhoods, security_zones, datacenters,
    policy_matrix (per environment), firewall device patterns, and
    app_dc_mappings in one payload."""
    from app.database import (
        get_neighbourhoods, get_security_zones, get_ngdc_datacenters,
        get_ngdc_prod_matrix, get_nonprod_matrix, get_preprod_matrix,
        get_app_dc_mappings, get_applications, get_firewall_devices,
        get_firewall_device_patterns, get_dc_vendor_map,
    )
    return {
        "neighbourhoods": await get_neighbourhoods(),
        "security_zones": await get_security_zones(),
        "datacenters": await get_ngdc_datacenters(),
        "policy_matrix": {
            "production": await get_ngdc_prod_matrix(),
            "non_production": await get_nonprod_matrix(),
            "pre_production": await get_preprod_matrix(),
        },
        "app_dc_mappings": await get_app_dc_mappings(),
        "applications": await get_applications(),
        "firewall_devices": await get_firewall_devices(),
        "firewall_device_patterns": await get_firewall_device_patterns(),
        "dc_vendor_map": await get_dc_vendor_map(),
    }
